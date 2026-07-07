#!/usr/bin/env python3
"""Anthropic Arena — question import script.

Converts a CSV (or Excel) question bank into the app's courses.json format,
and optionally uploads it straight to Firebase Firestore.

Usage:
  # 1) CSV -> assets/content/courses.json (bundled offline content)
  python import_questions.py questions.csv --out ../assets/content/courses.json

  # 2) CSV -> Firestore (live content, no app release needed)
  python import_questions.py questions.csv --upload --key serviceAccountKey.json

Excel (.xlsx) input works too if `openpyxl` is installed:
  pip install openpyxl
Firestore upload requires the Firebase Admin SDK:
  pip install firebase-admin

CSV columns (see questions_template.csv):
  course_id, course_title, course_tagline, course_description, course_order,
  course_color, level_id, level_title, level_order, level_topic, pass_mark,
  xp_per_correct, question_id, topic, question, option_a..option_d,
  correct (A/B/C/D), explanation, resource_title, resource_url
"""

import argparse
import csv
import json
import sys
from pathlib import Path

LETTER_TO_INDEX = {"A": 0, "B": 1, "C": 2, "D": 3}


def read_rows(path: Path):
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Excel input needs openpyxl: pip install openpyxl")
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        for raw in rows:
            if raw is None or all(v is None for v in raw):
                continue
            yield {header[i]: ("" if raw[i] is None else str(raw[i]).strip())
                   for i in range(min(len(header), len(raw)))}
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                yield {k.strip(): (v or "").strip() for k, v in row.items()}


def build(rows):
    courses = {}
    count = 0
    for i, row in enumerate(rows, start=2):  # header is line 1
        required = ["course_id", "level_id", "question_id", "question",
                    "option_a", "option_b", "option_c", "option_d", "correct"]
        missing = [c for c in required if not row.get(c)]
        if missing:
            sys.exit(f"Row {i}: missing required column(s): {', '.join(missing)}")
        correct = row["correct"].upper()
        if correct not in LETTER_TO_INDEX:
            sys.exit(f"Row {i}: 'correct' must be A, B, C or D (got {row['correct']!r})")

        course = courses.setdefault(row["course_id"], {
            "id": row["course_id"],
            "title": row.get("course_title", row["course_id"]),
            "tagline": row.get("course_tagline", ""),
            "description": row.get("course_description", ""),
            "order": int(row.get("course_order") or len(courses) + 1),
            "color": row.get("course_color", "#F5A623"),
            "_levels": {},
        })
        level = course["_levels"].setdefault(row["level_id"], {
            "id": row["level_id"],
            "title": row.get("level_title", row["level_id"]),
            "order": int(row.get("level_order") or len(course["_levels"]) + 1),
            "topic": row.get("level_topic", ""),
            "passMark": int(row.get("pass_mark") or 70),
            "xpPerCorrect": int(row.get("xp_per_correct") or 10),
            "questions": [],
        })
        question = {
            "id": row["question_id"],
            "topic": row.get("topic") or level["topic"],
            "question": row["question"],
            "options": [row["option_a"], row["option_b"], row["option_c"], row["option_d"]],
            "correctIndex": LETTER_TO_INDEX[correct],
            "explanation": row.get("explanation", ""),
        }
        if row.get("resource_url"):
            question["resource"] = {
                "title": row.get("resource_title", "Learn more"),
                "url": row["resource_url"],
            }
        level["questions"].append(question)
        count += 1

    out = {"version": 1, "courses": []}
    for course in sorted(courses.values(), key=lambda c: c["order"]):
        levels = sorted(course.pop("_levels").values(), key=lambda l: l["order"])
        course["levels"] = levels
        out["courses"].append(course)
    print(f"Parsed {count} questions across "
          f"{sum(len(c['levels']) for c in out['courses'])} levels "
          f"in {len(out['courses'])} courses.")
    return out


def upload(data, key_path):
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        sys.exit("Upload needs the Firebase Admin SDK: pip install firebase-admin")
    cred = credentials.Certificate(key_path)
    firebase_admin.initialize_app(cred)
    db = firestore.client()
    batch = db.batch()
    ops = 0
    for course in data["courses"]:
        course_doc = {k: v for k, v in course.items() if k != "levels"}
        batch.set(db.collection("courses").document(course["id"]), course_doc)
        ops += 1
        for level in course["levels"]:
            ref = (db.collection("courses").document(course["id"])
                     .collection("levels").document(level["id"]))
            batch.set(ref, level)
            ops += 1
            if ops >= 400:  # Firestore batch limit is 500 ops
                batch.commit()
                batch = db.batch()
                ops = 0
    if ops:
        batch.commit()
    print("Uploaded to Firestore.")


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", help="CSV or Excel file with questions")
    parser.add_argument("--out", default="courses.json", help="Output JSON path")
    parser.add_argument("--upload", action="store_true", help="Upload to Firestore")
    parser.add_argument("--key", help="Firebase service account key JSON (for --upload)")
    args = parser.parse_args()

    data = build(read_rows(Path(args.input)))

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {out_path}")

    if args.upload:
        if not args.key:
            sys.exit("--upload requires --key serviceAccountKey.json")
        upload(data, args.key)


if __name__ == "__main__":
    main()
